"""Add missing MILA Lounge 2025 shows from Act. Vs. For. Reporting into sched-baked.js.
Never overwrite existing performance fields on dates already present.
"""
from __future__ import annotations

import json
import re
from datetime import date, datetime
from pathlib import Path

import openpyxl

ROOT = Path(r"C:\Users\MatthiasLavenant\Documents\rdg-dj-dashboard")
BAKED = ROOT / "data" / "sched-baked.js"
XLSX = Path(
    r"c:\Users\MatthiasLavenant\mila-group.com\Riviera Dining Group Current - CONTROLLER"
    r"\2 - FP&A\1 - FP&A\10 - RDG\10 - Sales\ROI MINIMUM FLASH\MILA 2F"
    r"\MILA LOUNGE ROI DJ - 2025.xlsx"
)
VENUE = "MILA Lounge"


def _num(v):
    if v is None or v == "":
        return None
    try:
        n = float(v)
    except (TypeError, ValueError):
        return None
    if n != n:  # NaN
        return None
    return n


def _round4(v):
    n = _num(v)
    if n is None:
        return None
    return round(n, 4)


def load_excel():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Act. Vs. For. Reporting"]
    rows = {}
    for r in range(6, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if not isinstance(v, (datetime, date)):
            continue
        d = v.date() if isinstance(v, datetime) else v
        if d.year != 2025:
            continue
        dj = ws.cell(r, 3).value
        if isinstance(dj, (int, float)):
            dj = "" if float(dj) == 0 else str(int(dj) if float(dj) == int(dj) else dj)
        else:
            dj = (str(dj).strip() if dj is not None else "")
        if not dj or dj == "0":
            continue
        fee = _num(ws.cell(r, 4).value)
        bs_a = _num(ws.cell(r, 10).value)
        bs_m = _num(ws.cell(r, 11).value)
        roi_a = _round4(ws.cell(r, 14).value)
        roi_t = _num(ws.cell(r, 15).value)
        if roi_t is not None and roi_t <= 0:
            roi_t = None
        if fee is None and (bs_a is None or bs_a == 0):
            continue
        rows[d.isoformat()] = {
            "dj": dj,
            "fee": fee,
            "bs_a": bs_a,
            "bs_m": bs_m,
            "roi_a": roi_a,
            "roi_t": roi_t,
        }
    return rows


def parse_baked(text: str):
    m_sched = re.search(r"var SCHED\s*=\s*(\[.*?\]);\s*\nvar BS\s*=\s*", text, re.S)
    if not m_sched:
        raise SystemExit("Could not find SCHED array")
    sched = json.loads(m_sched.group(1))
    rest = text[m_sched.end() :]
    m_bs = re.match(r"(\[.*?\]);\s*\Z", rest, re.S)
    if not m_bs:
        # BS may be followed by more
        m_bs = re.match(r"(\[.*?\]);", rest, re.S)
    if not m_bs:
        raise SystemExit("Could not find BS array")
    bs = json.loads(m_bs.group(1))
    suffix = rest[m_bs.end() :]
    prefix = text[: m_sched.start()]
    return prefix, sched, bs, suffix


def existing_mila_dates(sched):
    out = set()
    for r in sched:
        if (r.get("venue") or r.get("v")) != VENUE:
            continue
        d = r.get("d")
        if not d or not str(d).startswith("2025-"):
            continue
        # treat as present if it has a DJ name (performance row)
        if (r.get("dj") or "").strip():
            out.add(d)
    return out


def make_sched_rec(ds, row):
    fee = row["fee"]
    bs_a = row["bs_a"]
    bs_m = row["bs_m"]
    roi_a = row["roi_a"]
    roi_t = row["roi_t"]
    beat = None
    status = "nd"
    if bs_a is not None and bs_m is not None:
        beat = 1 if bs_a >= bs_m else 0
        status = "beat" if beat else "miss"
    elif bs_a is not None:
        status = "beat"  # have actuals
        beat = None
    return {
        "venue": VENUE,
        "v": VENUE,
        "yr": 2025,
        "d": ds,
        "dj": row["dj"],
        "fee": fee,
        "cost": fee,
        "bs_a": bs_a,
        "bs_m": bs_m,
        "roi_a": roi_a,
        "roi_t": roi_t,
        "beat": beat,
        "ev": None,
        "_s": status,
    }


def make_bs_rec(ds, row):
    fee = row["fee"]
    bs_a = row["bs_a"]
    bs_m = row["bs_m"]
    beat = None
    if bs_a is not None and bs_m is not None:
        beat = 1 if bs_a >= bs_m else 0
    return {
        "venue": VENUE,
        "yr": 2025,
        "d": ds,
        "dj": row["dj"],
        "cost": fee,
        "bs_a": bs_a,
        "bs_m": bs_m,
        "roi_a": row["roi_a"],
        "roi_t": row["roi_t"] if row["roi_t"] is not None else 0,
        "beat": beat if beat is not None else 0,
    }


def main():
    text = BAKED.read_text(encoding="utf-8")
    prefix, sched, bs, suffix = parse_baked(text)
    excel = load_excel()
    present = existing_mila_dates(sched)
    bs_dates = {
        r["d"]
        for r in bs
        if r.get("venue") == VENUE and str(r.get("d", "")).startswith("2025-")
    }

    to_add = sorted(d for d in excel if d not in present)
    print("existing MILA 2025 dates", len(present))
    print("excel usable rows", len(excel))
    print("adding", len(to_add), "SCHED rows:", to_add)

    added_sched = 0
    for ds in to_add:
        sched.append(make_sched_rec(ds, excel[ds]))
        added_sched += 1

    added_bs = 0
    for ds in to_add:
        if ds in bs_dates:
            continue
        if excel[ds].get("bs_a") is None:
            continue
        bs.append(make_bs_rec(ds, excel[ds]))
        added_bs += 1

    # Stable-ish order: keep original then append (append-only is safest)
    out = (
        prefix
        + "var SCHED = "
        + json.dumps(sched, separators=(",", ":"), ensure_ascii=False)
        + ";\n"
        + "var BS    = "
        + json.dumps(bs, separators=(",", ":"), ensure_ascii=False)
        + ";"
        + suffix
    )
    BAKED.write_text(out, encoding="utf-8")
    print("wrote", BAKED)
    print("added_sched", added_sched, "added_bs", added_bs)

    # Verify no existing Aug 2 performance changed
    for r in sched:
        if r.get("d") == "2025-08-02" and (r.get("venue") or r.get("v")) == VENUE:
            print("verify Aug2 still", r.get("dj"), r.get("bs_a"), r.get("fee"))


if __name__ == "__main__":
    main()
